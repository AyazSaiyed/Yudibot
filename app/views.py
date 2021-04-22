# -*- encoding: utf-8 -*-

'''Developer   - Ayaz Saiyed M.
   Last update - April 22, 2021
   Department  - AI/ML
'''

from django.contrib.auth.decorators import login_required
from django.shortcuts import render, get_object_or_404, redirect
from django.template import loader
from django.http import HttpResponse
from django import template

from django.shortcuts import render
from django.http import HttpResponse
from django.views.decorators.csrf import csrf_exempt
from django.http import Http404
from django.http import HttpResponse, JsonResponse
import smtplib
from django.http import Http404
from pydialogflow_fulfillment import DialogflowResponse, DialogflowRequest, SimpleResponse, Suggestions, LinkOutSuggestion
from django.http import HttpResponse, JsonResponse
from library.df_response_lib import *
from library.facebook_template_lib import *
import json
import requests
from .models import ClientEnquiries,JobApplicants
# import facebook
import time
import random
from django.shortcuts import redirect
from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl import load_workbook

from datetime import datetime
import smtplib
# import request
import datetime
import pickle
import os.path
from googleapiclient.discovery import build
from google_auth_oauthlib.flow import InstalledAppFlow
from google.auth.transport.requests import Request
from datetime import timedelta
from oauth2client.service_account import ServiceAccountCredentials
SCOPES = ['https://www.googleapis.com/auth/calendar']
# import request

import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders

@login_required(login_url="/login/")
def index(request):
    context = ClientEnquiries.objects.all()
    print(" Context details ",context)
    print("length", len(context))
    totalenquries = len(context)
    # html_template = loader.get_template('tables.html')
    # return HttpResponse(html_template.render(context, request))
    return render(request,'tables.html',{'clientenquiries':context,'totalenquries':totalenquries})


@login_required(login_url="/login/")
def jobapplicants(request):
    context = JobApplicants.objects.all()
    print(" Context details ",context)
    print("length", len(context))
    totalcandidates = len(context)
    # html_template = loader.get_template('tables.html')
    # return HttpResponse(html_template.render(context, request))
    return render(request,'jobapplicants.html',{'jobapplicants':context,'totalcandidates':totalcandidates})

@login_required(login_url="/login/")
def pages(request):
    context = {}
    # All resource paths end in .html.
    # Pick out the html file name from the url. And load that template.
    try:

        load_template = request.path.split('/')[-1]
        html_template = loader.get_template( load_template )
        return HttpResponse(html_template.render(context, request))

    except template.TemplateDoesNotExist:

        html_template = loader.get_template( 'error-404.html' )
        return HttpResponse(html_template.render(context, request))

    except:

        html_template = loader.get_template( 'error-500.html' )
        return HttpResponse(html_template.render(context, request))




# https://yudizsolutions.pythonanywhere.com/bot/
# https://myaccount.google.com/lesssecureapps?
# https://accounts.google.com/DisplayUnlockCaptcha




fromaddr = "yudizblog@gmail.com"
toaddr = "saiyedayaz99@gmail.com"

    # instance of MIMEMultipart
msg = MIMEMultipart()
    # storing the senders email address
msg['From'] = fromaddr
    # storing the receivers email address
msg['To'] = toaddr
    # storing the subject
msg['Subject'] = "New update from Yudibot"
msg['Subject1'] = "Client Enquiry"

    # string to store the body of the mail
body = "Please review the attached sheet"
msg.attach(MIMEText(body, 'plain'))
s = smtplib.SMTP('smtp.gmail.com', 587)

    # start TLS for security
s.starttls()
    # Authentication
s.login(fromaddr, "Yudiz@1527")
    # open the file to be sent
    # creates SMTP session
    # open the file to be sent
    # creates SMTP session


a = []
uemail = []
@csrf_exempt
def index_function(request):
   if request.method == "POST":

        print("Method ",request.method)

        if request.body:
            req = json.loads(request.body)
            dialogflow_request = DialogflowRequest(request.body)
            action = req.get('queryResult').get('action')
            print("Action is ",action)
            if dialogflow_request.get_intent_displayName() == "Default_Welcome_Intent_Join":
                print("-----------------------")
                print("In Default_Welcome_Intent_Join ")
                dialogflow_response = DialogflowResponse("Okay,now please select the field in which you want to work with")
                dialogflow_response.add(SimpleResponse("a","Please select the field in which you want to work with"))
                dialogflow_response.add(Suggestions(["Android Developer","Python Developer","Designer","Mean Stack Developer","Game Developer","Game Designer","Web Development"]))
                req = json.loads(request.body)
                action = req.get('queryResult').get('action')
                x = action

                print("Join Yudiz Intent")
                response = dialogflow_response.get_final_response()





            if dialogflow_request.get_intent_displayName() == "Default Welcome Intent -Sales - Project Development - Email - custom":
                print("Inside Default Welcome Intent -Sales - Project Development - Email - custom")
                fb = facebook_response()
                texts = ['Perfect, we have received your details successfully. One of our executive will get back to you after reviewing your details. Thanks for applying in Yudiz 🤝']
                fb_text_replies = fb.text_response(texts)
                action = req.get('queryResult').get('action')
                print("details ",action)
                usersname = req.get('queryResult').get('parameters').get('nameemail')
                email = req.get('queryResult').get('parameters').get('email')
                requirement = req.get('queryResult').get('parameters').get('projectrequirement')
                phone = req.get('queryResult').get('parameters').get('phone')
                print("usersname ",usersname)


                fulfillmentText = 'Suggestion chips Response from webhook'
                ff_response = fulfillment_response()
                ff_text = ff_response.fulfillment_text(fulfillmentText)

                title = ""+usersname+", would you like to continue conversation ?"
                replies= ["🔙 Main Menu 🔖","Contact Us ☎️","About Us 🏢"," Why Yudiz ?","Our Services ⚙️"]
                fb_quick_replies = fb.quick_replies(title, replies)


                ff_messages = ff_response.fulfillment_messages([fb_text_replies,fb_quick_replies])
                reply = ff_response.main_response(ff_text, ff_messages)

                workbook = Workbook()
                # workbook = load_workbook(filename="/home/YudiBot/argon-dashboard-django-master/ClientEnquiry.xlsx")
                workbook = load_workbook(filename="./ClientEnquiry.xlsx")

                sheet = workbook.active
                strrequirement = str(requirement)
                # Details of User
                rows = [
                            [usersname,strrequirement,email,phone]
                       ]

                for row in rows:
                    sheet.append(row)


                # workbook.save(filename="/home/YudiBot/argon-dashboard-django-master/ClientEnquiry.xlsx")
                workbook.save(filename="./ClientEnquiry.xlsx")


                for col in sheet.columns:
                    max_length = 0
                    column = col[0].column # Get the column name
                    for cell in col:
                        if cell.coordinate in sheet.merged_cells: # not check merge_cells
                            continue
                        try: # Necessary to avoid error on empty cells
                            if len(str(cell.value)) > max_length:
                                max_length = len(cell.value)
                        except:
                            pass
                    adjusted_width = int((max_length + 2) * 1.2)
                    print("adjusted_width",adjusted_width)
                    sheet.column_dimensions[col[0].column_letter].width = adjusted_width
                workbook.save(filename="./ClientEnquiry.xlsx")
                # import uuid
                # uid = uuid.uuid4()
                ClientEnquiries.objects.create(username=usersname,enquriy=strrequirement,email=email,phone=phone)
                print(" Data Saved ")
                action = req.get('queryResult').get('action')
                x = action
                filename = "ClientEnquiry.xlsx"
                attachment = open("./ClientEnquiry.xlsx", "rb")
                try:
                    p = MIMEBase('application', 'octet-stream')
                    p.set_payload((attachment).read())
                    encoders.encode_base64(p)
                    p.add_header('Content-Disposition', "attachment; filename= %s" % filename)
                        # attach the instance 'p' to instance 'msg'
                    msg.attach(p)
                    text = msg.as_string()
                    toaddrSales = 'saiyedayaz9@gmail.com'
                    s.sendmail(fromaddr, toaddrSales, text)
                    print(" Client Enquiry has been sent successfully ")
                    s.quit()
                    return JsonResponse(reply, safe=False)

                except:
                    print(" Email Has Been Sent ")


                return JsonResponse(reply, safe=False)




            if dialogflow_request.get_intent_displayName() == "Default_Welcome_Intent_Join_AndroidEmail":
                print("Inside Default_Welcome_Intent_Join_AndroidEmail")
                fb = facebook_response()
                texts = ['Greetings, we have received your details successfully. We will review your details and our HR will get back to you soon. Thanks for applying in Yudiz 🤝']
                fb_text_replies = fb.text_response(texts)
                action = req.get('queryResult').get('action')
                print("details ",action)
                usersname = req.get('queryResult').get('parameters').get('nameemail')
                email = req.get('queryResult').get('parameters').get('email')
                requirement = req.get('queryResult').get('parameters').get('projectrequirement')
                print("usersname ",usersname)
                fulfillmentText = 'Suggestion chips Response from webhook'
                ff_response = fulfillment_response()
                ff_text = ff_response.fulfillment_text(fulfillmentText)
                title = ""+usersname+", would you like to continue conversation ?"
                replies= ["Main Menu 🔖","Contact Us ☎️","About Us 🏢"," Why Yudiz ?","Our Services ⚙️","Play Game 🎰"]
                fb_quick_replies = fb.quick_replies(title, replies)
                ff_messages = ff_response.fulfillment_messages([fb_text_replies,fb_quick_replies])
                reply = ff_response.main_response(ff_text, ff_messages)
                return JsonResponse(reply, safe=False)


            if dialogflow_request.get_intent_displayName() == "Default_Welcome_Intent_Join_AndroidEmail -Resume":
                fb = facebook_response()
                usersname = req.get('queryResult').get('parameters').get('usersname')
                usersemail = req.get('queryResult').get('parameters').get('usersemail')
                applyingfor = req.get('queryResult').get('parameters').get('applyingfor')
                totalexperience = req.get('queryResult').get('parameters').get('totalexperience')
                url = req.get('queryResult').get('parameters').get('url')
                print("totalexperience",totalexperience)

                texptemp = req.get('queryResult').get('parameters')
                print("texptemp",texptemp)
                texts = ['Great, we have received your resume for position of '+applyingfor+', after reviewing it, our HR team will connect with you soon ☎️']
                fb_text_replies = fb.text_response(texts)
                # print("All details ",req.get('originalDetectIntentRequest').get('payload').get('data').get('message').get('attachments'))
                AttachedResume = req.get('originalDetectIntentRequest').get('payload').get('data').get('message').get('attachments')[0].get('payload').get('url')
                print("AttachedResume",AttachedResume)


                fulfillmentText = 'Suggestion chips Response from webhook'
                ff_response = fulfillment_response()
                ff_text = ff_response.fulfillment_text(fulfillmentText)

                title = ""+usersname+" ,would you like to continue conversation ?"
                replies= ["🔙 Main Menu","Contact Us ☎️","About Us 🏢"," Why Yudiz ?","Our Services⚙️"]
                fb_quick_replies = fb.quick_replies(title, replies)
                ff_messages = ff_response.fulfillment_messages([fb_text_replies,fb_quick_replies])
                reply = ff_response.main_response(ff_text, ff_messages)

# --------------------------
                #Excel logs stuff
                totalexperienceStr = str(totalexperience)
                workbook = Workbook()
                workbook = load_workbook(filename="./JobApplicantsList.xlsx")
                sheet = workbook.active
                # Details of User
                rows = [
                            [usersname,usersemail,applyingfor,totalexperienceStr,AttachedResume]
                       ]

                for row in rows:
                    sheet.append(row)

                workbook.save(filename="./JobApplicantsList.xlsx")
                for col in sheet.columns:
                    max_length = 0
                    column = col[0].column # Get the column name
                    for cell in col:
                        if cell.coordinate in sheet.merged_cells: # not check merge_cells
                            continue
                        try: # Necessary to avoid error on empty cells
                            if len(str(cell.value)) > max_length:
                                max_length = len(cell.value)
                        except:
                            pass
                    adjusted_width = int((max_length + 2) * 1.2)
                    print("adjusted_width",adjusted_width)
                    sheet.column_dimensions[col[0].column_letter].width = adjusted_width
                workbook.save(filename="./JobApplicantsList.xlsx")

                print(" Data Saved ")
                action = req.get('queryResult').get('action')
                x = action
                filename = "JobApplicantsList.xlsx"
                attachment = open("./JobApplicantsList.xlsx", "rb")
                try:
                    p = MIMEBase('application', 'octet-stream')
                    p.set_payload((attachment).read())
                    encoders.encode_base64(p)
                    p.add_header('Content-Disposition', "attachment; filename= %s" % filename)
                        # attach the instance 'p' to instance 'msg'
                    msg.attach(p)
                    text = msg.as_string()
                    toaddrHR = 'hr@yudiz.com'
                    s.sendmail(fromaddr, toaddr, text)
                    print(" Email Has Been Sent ")
                    s.quit()
                    return JsonResponse(reply, safe=False)

                except:
                    print(" Email Has Been Sent ")

                return JsonResponse(reply, safe=False)



#for resume via link
            if dialogflow_request.get_intent_displayName() == "FetchResumeOption(DropLink) - custom":
                fb = facebook_response()

                usersname = req.get('queryResult').get('parameters').get('username')
                email = req.get('queryResult').get('parameters').get('email')
                applyingfor = req.get('queryResult').get('parameters').get('jobprofile')
                totalexperience = req.get('queryResult').get('parameters').get('experience')
                url = req.get('queryResult').get('parameters').get('url')

                texts = ['Great, we have received your resume, after reviewing it our HR team will connect with you soon ☎️ Thank you for applying in Yudiz.']
                fb_text_replies = fb.text_response(texts)

                fulfillmentText = 'Suggestion chips Response from webhook'
                ff_response = fulfillment_response()
                ff_text = ff_response.fulfillment_text(fulfillmentText)
                print("usersname ",usersname)
                title = usersname+",would you like to continue conversation ?"
                replies= ["🔙 Main Menu","Contact Us ☎️","About Us 🏢"," Why Yudiz ?","Our Services⚙️"]
                fb_quick_replies = fb.quick_replies(title, replies)

                ff_messages = ff_response.fulfillment_messages([fb_text_replies,fb_quick_replies])
                reply = ff_response.main_response(ff_text, ff_messages)

                #Excel logs stuff
                totalexperienceStr = str(totalexperience)
                workbook = Workbook()
                workbook = load_workbook(filename="./JobApplicantsList.xlsx")
                sheet = workbook.active
                # Details of User
                rows = [
                            [usersname,applyingfor,totalexperienceStr,email,url]
                       ]

                for row in rows:
                    sheet.append(row)

                JobApplicants.objects.create(candidatename=usersname,appliedfor=applyingfor,email=email,experience=totalexperienceStr,resumeurl=url)
                print("Details saved in JobApplicants")
                workbook.save(filename="./JobApplicantsList.xlsx")
                for col in sheet.columns:
                    max_length = 0
                    column = col[0].column # Get the column name
                    for cell in col:
                        if cell.coordinate in sheet.merged_cells: # not check merge_cells
                            continue
                        try: # Necessary to avoid error on empty cells
                            if len(str(cell.value)) > max_length:
                                max_length = len(cell.value)
                        except:
                            pass
                    adjusted_width = int((max_length + 2) * 1.2)
                    print("adjusted_width",adjusted_width)
                    sheet.column_dimensions[col[0].column_letter].width = adjusted_width
                workbook.save(filename="./JobApplicantsList.xlsx")

                print(" Data Saved ")
                action = req.get('queryResult').get('action')
                x = action
                filename = "JobApplicantsList.xlsx"
                attachment = open("./JobApplicantsList.xlsx", "rb")
                try:
                    p = MIMEBase('application', 'octet-stream')
                    p.set_payload((attachment).read())
                    encoders.encode_base64(p)
                    p.add_header('Content-Disposition', "attachment; filename= %s" % filename)
                        # attach the instance 'p' to instance 'msg'
                    msg.attach(p)
                    text = msg.as_string()
                    toaddrHR = 'hr@yudiz.com'
                    toaddrCTO = 'pankit@yudiz.com'
                    s.sendmail(fromaddr, toaddr, text)
                    print(" Email Has Been Sent ")
                    s.quit()
                    return JsonResponse(reply, safe=False)

                except:
                    print(" Email Not Sent ")

                return JsonResponse(reply, safe=False)


            if dialogflow_request.get_intent_displayName() == "FetchResumeOption(dropResume) - custom":
                fb = facebook_response()

                usersname = req.get('queryResult').get('parameters').get('username')
                usersemail = req.get('queryResult').get('parameters').get('email')
                applyingfor = req.get('queryResult').get('parameters').get('jobprofile')
                totalexperience = req.get('queryResult').get('parameters').get('experience')
                # url = req.get('queryResult').get('parameters').get('url')
                # print(" Resume URL ",url)
                print("totalexperience",totalexperience)

                texptemp = req.get('queryResult').get('parameters')
                print("texptemp",texptemp)
                texts = ['Great, we have received your resume, after reviewing it our HR team will connect with you soon ☎️']
                fb_text_replies = fb.text_response(texts)
                print("All details ",req.get('originalDetectIntentRequest').get('payload').get('data').get('message').get('attachments'))
                print("RRRreq.get('originalDetectIntentRequest').get('payload').get('data')",req.get('originalDetectIntentRequest').get('payload').get('data').get('message').get('attachments'))
                try:
                    AttachedResume = req.get('originalDetectIntentRequest').get('payload').get('data').get('message').get('attachments')[0].get('payload').get('url')
                    print("AttachedResume",AttachedResume)
                    fulfillmentText = 'Suggestion chips Response from webhook'
                    ff_response = fulfillment_response()
                    ff_text = ff_response.fulfillment_text(fulfillmentText)

                    title = ""+usersname+", would you like to continue conversation ?"
                    replies= ["🔙 Main Menu","Contact Us ☎️","About Us 🏢"," Why Yudiz ?","Our Services⚙️"]
                    fb_quick_replies = fb.quick_replies(title, replies)


                    ff_messages = ff_response.fulfillment_messages([fb_text_replies,fb_quick_replies])
                    reply = ff_response.main_response(ff_text, ff_messages)

    # --------------------------
                    #Excel logs stuff
                    totalexperienceStr = str(totalexperience)
                    workbook = Workbook()
                    workbook = load_workbook(filename="./JobApplicantsList.xlsx")
                    sheet = workbook.active
                    # Details of User
                    rows = [
                                [usersname,usersemail,applyingfor,totalexperienceStr,AttachedResume]
                           ]
                    JobApplicants.objects.create(candidatename=usersname,appliedfor=applyingfor,email=usersemail,experience=totalexperienceStr,resumeurl=AttachedResume)
                    print("Details saved in JobApplicants")

                    for row in rows:
                        sheet.append(row)


                    workbook.save(filename="./JobApplicantsList.xlsx")
                    for col in sheet.columns:
                        max_length = 0
                        column = col[0].column # Get the column name
                        for cell in col:
                            if cell.coordinate in sheet.merged_cells: # not check merge_cells
                                continue
                            try: # Necessary to avoid error on empty cells
                                if len(str(cell.value)) > max_length:
                                    max_length = len(cell.value)
                            except:
                                pass
                        adjusted_width = int((max_length + 2) * 1.2)
                        print("adjusted_width",adjusted_width)
                        sheet.column_dimensions[col[0].column_letter].width = adjusted_width
                    workbook.save(filename="./JobApplicantsList.xlsx")

                    print(" Data Saved ")
                    action = req.get('queryResult').get('action')
                    x = action
                    filename = "JobApplicantsList.xlsx"
                    attachment = open("./JobApplicantsList.xlsx", "rb")
                    try:
                        p = MIMEBase('application', 'octet-stream')
                        p.set_payload((attachment).read())
                        encoders.encode_base64(p)
                        p.add_header('Content-Disposition', "attachment; filename= %s" % filename)
                            # attach the instance 'p' to instance 'msg'
                        msg.attach(p)
                        text = msg.as_string()
                        toaddrSales = 'saiyedayaz9@gmail.com'
                        s.sendmail(fromaddr, toaddrSales, text)
                        print(" Resume received successfully ")
                        s.quit()
                        return JsonResponse(reply, safe=False)
                    except:
                        print(" Email Has not been sent ")
                    return JsonResponse(reply, safe=False)
                except:
                    AttachedResume = "No Resume Found"
                    print("AttachedResume",AttachedResume)
                    fulfillmentText = 'Suggestion chips Response from webhook'
                    ff_response = fulfillment_response()
                    ff_text = ff_response.fulfillment_text(fulfillmentText)

                    title = ""+usersname+", would you like to continue conversation ?"
                    replies= ["🔙 Main Menu","Contact Us ☎️","About Us 🏢"," Why Yudiz ?","Our Services⚙️"]
                    fb_quick_replies = fb.quick_replies(title, replies)


                    ff_messages = ff_response.fulfillment_messages([fb_text_replies,fb_quick_replies])
                    reply = ff_response.main_response(ff_text, ff_messages)

    # --------------------------
                    #Excel logs stuff
                    totalexperienceStr = str(totalexperience)
                    workbook = Workbook()
                    workbook = load_workbook(filename="./JobApplicantsList.xlsx")
                    sheet = workbook.active
                    # Details of User
                    rows = [
                                [usersname,usersemail,applyingfor,totalexperienceStr,AttachedResume]
                           ]

                    for row in rows:
                        sheet.append(row)

                    JobApplicants.objects.create(candidatename=usersname,appliedfor=applyingfor,email=usersemail,experience=totalexperienceStr,resumeurl=AttachedResume)
                    print("Details saved in JobApplicants")
                    workbook.save(filename="./JobApplicantsList.xlsx")
                    for col in sheet.columns:
                        max_length = 0
                        column = col[0].column # Get the column name
                        for cell in col:
                            if cell.coordinate in sheet.merged_cells: # not check merge_cells
                                continue
                            try: # Necessary to avoid error on empty cells
                                if len(str(cell.value)) > max_length:
                                    max_length = len(cell.value)
                            except:
                                pass
                        adjusted_width = int((max_length + 2) * 1.2)
                        print("adjusted_width",adjusted_width)
                        sheet.column_dimensions[col[0].column_letter].width = adjusted_width
                    workbook.save(filename="./JobApplicantsList.xlsx")

                    print(" Data Saved ")
                    action = req.get('queryResult').get('action')
                    x = action
                    filename = "JobApplicantsList.xlsx"
                    attachment = open("./JobApplicantsList.xlsx", "rb")
                    try:
                        p = MIMEBase('application', 'octet-stream')
                        p.set_payload((attachment).read())
                        encoders.encode_base64(p)
                        p.add_header('Content-Disposition', "attachment; filename= %s" % filename)
                            # attach the instance 'p' to instance 'msg'
                        msg.attach(p)
                        text = msg.as_string()
                        toaddrSales = 'saiyedayaz9@gmail.com'
                        s.sendmail(fromaddr, toaddrSales, text)
                        print(" Resume received successfully ")
                        s.quit()
                        return JsonResponse(reply, safe=False)
                    except:
                        print(" Email Has not been sent ")
                    return JsonResponse(reply, safe=False)
                # AttachedResume = url




            if dialogflow_request.get_intent_displayName() == "Default Welcome Intent -Sales - Project Development - custom":
                print("-----------------------")

                print("In project development 4")

                req = json.loads(request.body)
                action = req.get('queryResult').get('action')
                action1 = req.get('queryResult').get('queryText')
                # print("Pdata",Pdata)

                username = action1
                b = username

                x = action
                print(" ")
                print("-------- All",action)
                print("-------- username",b)
                print(" ")
                ui = "ProjectDevelopment"
                request.session['usersinterest'] = ui
                bb = str(b)
                request.session['username'] = bb
                print("-------------------------------------------------")

                print("-------------------------------------------------")
                print("Username is",request.session.get('username'))
                uu = request.session.get('username')
                print("Users Interest is",request.session.get('usersinterest'))
                ii = request.session.get('usersinterest')


                print("-------------------------------------------------")

                print("-------------------------------------------------")
                a.append(uu)
                a.append(ii)
                print("Array is",a)



            if dialogflow_request.get_intent_displayName() == "Default Welcome Intent -Sales - Project Development - Email":
                dialogflow_response = DialogflowResponse("a","Great ! We are having a openings in Python Development")
                # print("username",b)
                print("-----------------------")

                print("In Email 4")
                # print(" Array till now before email ---- ",Pdata)

                req = json.loads(request.body)
                action = req.get('queryResult')
                action1 = req.get('queryResult').get('queryText')
                action2 = req.get('queryResult').get('person.original')

                # dialogflow_response = DialogflowResponse("Great"+b+" All Set.Let us know which field you are interested ")

                dialogflow_response.add(SimpleResponse("","All Set. Let us know which field you are interested"))
                dialogflow_response.add(Suggestions(["Create Game","Build Website","Build Android Application","Build IOS Application"]))

                email = action1
                # print(" Array till now after email ---- ",Pdata)
                # Pdata.append(email)

                # v = str(action1)
                # print("Username is ",action1.person)
                x = action1
                xx = str(x)
                print(" ")
                # print("-------- All",action)
                print("-------- Email",action1)
                print("-------- email",action2)
                print(" ")
                print(" ")
                request.session['emailofuserq'] = xx
                y = request.session['emailofuserq']
                # mail()

                print("I just set the session value to ",y)
                print("-------------------------------------------------")
                print("Users Email is",request.session.get('emailofuserq'))
                ee = request.session.get('emailofuserq')
                print("-------------------------------------------------")
                # s = smtplib.SMTP('smtp.gmail.com', 587)
                # s.starttls()
                # Authentication
                ee = request.session.get('emailofuserq')
                a.append(ee)
                uemail.append(ee)
                response = dialogflow_response.get_final_response()



            if dialogflow_request.get_intent_displayName() == "Default Welcome Intent -Sales - Project Development - Android Interest":
                dialogflow_response = DialogflowResponse("a","Great ! User Interest")

                req = json.loads(request.body)
                action = req.get('queryResult')
                action1 = req.get('queryResult').get('queryText')
                # action2 = req.get('queryResult').get('person.original')

                # dialogflow_response = DialogflowResponse("Great"+b+" All Set.Let us know which field you are interested ")
                dialogflow_response.add(SimpleResponse("","Glad to let you know that we hold a rank in developing Android Applications"))
                dialogflow_response.add(Suggestions(["Why Yudiz","Our Developed Applications", "Get Quote"]))

                # dialogflow_response.add(Suggestions(["Create Game","Build Website","Build Android Application","Build IOS Application"]))
                # print("- - - -- - - -- - - -- Array : ",Pdata)
                usersinterest1 = action1

                field = "Android Development"
                # print("username",request.session.get('username'))
                # request.session['usersinterestfield'] = usersinterest1

                a.append(field)
                print(" - - - - - ")
                print("Array is",a)
                mail(request)
                response = dialogflow_response.get_final_response()



            if dialogflow_request.get_intent_displayName() == "schedulemeeting":
                req = json.loads(request.body)
                print(" Inside ")
                import datefinder

                # if os.path.exists('/home/YudiBot/argon-dashboard-django-master/token.pkl'):
                #     print("token found")
                #     with open('/home/YudiBot/argon-dashboard-django-master/token.pkl', 'rb') as token:
                #         creds = pickle.load(token, encoding='latin1')


                from datetime import timedelta

                import pytz
                from googleapiclient.discovery import build
                from oauth2client.service_account import ServiceAccountCredentials

                service_account_email = "yudizchatbot@yudizbot-abfjvm.iam.gserviceaccount.com"
                SCOPES = ["https://www.googleapis.com/auth/calendar"]
                creds = ServiceAccountCredentials.from_json_keyfile_name(
                    filename="yudizbot-abfjvm-cf29736bf66c.json", scopes=SCOPES
                )

                FinalSlotsList = []
                bookedslots1 = []
                FinalAvailableSlots = []
                username = req.get('queryResult').get('parameters').get('username')
                phone = req.get('queryResult').get('parameters').get('phone')
                fb = facebook_response()

                service = build('calendar', 'v3', credentials=creds)

                event_summary = 'Yudiz Solutions - Appointment via Chatbot'
    #             # event_location = 'H7Na 5H9aaaaaa'
                event_description = 'Scheduled appointment via Yudibot, Client name is '+username+', Contact no. is '+phone+'. Thank you !'
                TIMEZONE_OFFSET = 'Asia/Kolkata'

                req = json.loads(request.body)
                parameters = req['queryResult']['parameters']
                print(json.dumps(parameters, indent=4))
                start_date = parameters['date'].split('T')[0] + 'T' + parameters['time'].split('T')[1].split('+')[0] +'+05:30'
                print("Start Date",start_date)
                zStart = list(datefinder.find_dates(start_date))[0]
                xEnd = str(zStart + timedelta(minutes=20))


                xEnd1 = list(datefinder.find_dates(xEnd))[0]

                print(" Final Start and End timings")
                print(zStart)


                FinalStartTime = zStart.strftime("%Y-%m-%dT%H:%M:%S%Z")+'+05:30'
                FinalEndTime = xEnd1.strftime("%Y-%m-%dT%H:%M:%S%Z")+'+05:30'
                # zStart1 = zStart
                print("====>",FinalStartTime)
                # print(zStart)
                print("====>",FinalEndTime)

                # events_result = service.events().list(calendarId='fgc3l1cir3no88opvbca9ej0hc@group.calendar.google.com', timeMin=FinalStartTime,
                #                           maxResults=20, singleEvents=True,timeZone='Asia/Kolkata',
                #                           orderBy='startTime').execute()
                events_result = service.events().list(calendarId='cl0kjg2k6odsmt8kgc1d8nmrso@group.calendar.google.com', timeMin=FinalStartTime,
                                          maxResults=20, singleEvents=True,timeZone='Asia/Kolkata',
                                          orderBy='startTime').execute()
                                        #
                events = events_result.get('items',[])

                calendar_id = 'fgc3l1cir3no88opvbca9ej0hc@group.calendar.google.com'
                events = events_result.get('items',[])
                body = {
                    "timeMin": FinalStartTime,
                    "timeMax": FinalEndTime,
                    "timeZone": 'Asia/Kolkata',
                    "items": [{"id": calendar_id}]
                }
                freebusy_result = service.freebusy().query(body=body).execute()
                slots = freebusy_result.get('calendars', {}).get(calendar_id, [])
                print("slots ",len(slots['busy']))
                print("slots['busy']",slots['busy'])
                if len(slots['busy']) >= 1:
                    print(" non empty ")
                    print("Found Busy Slots")
                    # dialogflow_response = DialogflowResponse("Slots Unavailable")
                    # dialogflow_response.add(SimpleResponse("Slots Unavailable","Slots Unavailable"))

                    fulfillmentText = 'Suggestion chips Response from webhook'
                    ff_response = fulfillment_response()
                    ff_text = ff_response.fulfillment_text(fulfillmentText)

                    title = ""+username+", sorry but we already have appointment at the selected date and time , please try with different time "
                    replies= ["Book for tomorrow?","🔙 Main Menu","Book on friday?"," Why Yudiz ?","Call Sales📞"]
                    fb_quick_replies = fb.quick_replies(title, replies)


                    ff_messages = ff_response.fulfillment_messages([fb_quick_replies])
                    reply = ff_response.main_response(ff_text, ff_messages)
                    return JsonResponse(reply, safe=False)

                else:
                    texts = ['Greetings, '+username+' we have successfully scheduled your appointment 📮 One of our representative will contact you soon. Thank you for showing interest in Yudiz 🤝']
                    fb_text_replies = fb.text_response(texts)
                    print(" empty ")
                    event = {
                        'summary': event_summary,
                        # 'location': event_location,
                        'description': event_description,
                        'start': {
                            'dateTime': FinalStartTime,
                            'timeZone': 'Asia/Kolkata',
                        },
                        'end': {
                            'dateTime': FinalEndTime,
                            'timeZone': 'Asia/Kolkata',
                        },
                        'reminders': {
                            'useDefault': True,
                        },
                        'attendees': {
                            'email': 'yudizblog@gmail.com',
                        },
                    }
                    event = service.events().insert(calendarId='cl0kjg2k6odsmt8kgc1d8nmrso@group.calendar.google.com', body=event, sendUpdates='all').execute()
                    print("Appointment Booked Man")
                    print("No busy slots in the specified timeframe.")
                    # dialogflow_response = DialogflowResponse("Your appointment had been scheduled succcessfully")
                    # dialogflow_response.add(SimpleResponse("Your appointment had been scheduled succcessfully","Your appointment had been scheduled succcessfully"))

                    fulfillmentText = 'Suggestion chips Response from webhook'
                    ff_response = fulfillment_response()
                    ff_text = ff_response.fulfillment_text(fulfillmentText)

                    title = "How else we can help you ?"
                    replies= ["🔙 Main Menu 🔖","Contact Us ☎️","About Us 🏢"," Why Yudiz ?","Our Services"]
                    fb_quick_replies = fb.quick_replies(title, replies)


                    ff_messages = ff_response.fulfillment_messages([fb_text_replies,fb_quick_replies])
                    reply = ff_response.main_response(ff_text, ff_messages)
                    return JsonResponse(reply, safe=False)
                # response = dialogflow_response.get_final_response()
                # return JsonResponse(reply, safe=False)



            if dialogflow_request.get_intent_displayName() == "Default Welcome Intent -Sales - Project discussion":
                dialogflow_response = DialogflowResponse(" We are always to open to discuss solutions. ")
                dialogflow_response.add(SimpleResponse("a","We are always to open to discuss solutions."))
                print("Project Discussion Intent")
                req = json.loads(request.body)
                action = req.get('queryResult').get('action')
                action1 = req.get('queryResult')
                print("A ----- ",action1)

                x = action
                # print("user entered",x)
                response = dialogflow_response.get_final_response()
            else :
                response = {
                "error" : "1",
                "message" : "An error occurred."
                }
            return HttpResponse(response, content_type='application/json; charset=utf-8')
        else:
            raise Http404()


def temp(request):
    return render(request,'bot/yudizbot.html')



# ----------------------------------

'''Developer   - Ayaz Saiyed M.
   Last update - April 8, 2021
   Department  - AI/ML
'''